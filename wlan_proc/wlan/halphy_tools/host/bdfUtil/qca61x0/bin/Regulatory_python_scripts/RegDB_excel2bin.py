#import timeit # very important to calculate time complexities
from sys import argv
from openpyxl import load_workbook as lw

import subprocess

#script, NvData = argv

def is_number(s):
    try:
        float(s) # for int, long and float
    except ValueError:
        try:
            complex(s) # for complex
        except ValueError:
            return False

    return True

def clear_contents_of_column_in_sheet(ws,col):
    for row in ws[col+str(2)+':'+col+str(ws.max_row)]:
        for cell in row:
            cell.value = None

def put_zeroes_column_in_sheet(ws,col):
    for row in ws[col+str(2)+':'+col+str(ws.max_row)]:
        for cell in row:
            cell.value = 0
            
def combine_sheets(wi,wo,wbo,find_str):
    wi = wbo.get_sheet_by_name(wi)   
    wo = wbo.get_sheet_by_name(wo) 

    for i in range(1,wo.max_row+1):
        if (str(wo.cell(row=i,column=1).value).find(find_str) != -1):
            break

    for j in range(1,wi.max_row):
        wo.cell(row=i+(j-1),column=2).value = wi.cell(row=j+1,column=2).value


#wbi = lw("Regulatory_BDF_Data.xlsx")
wbi = lw(filename = "Regulatory_BDF_In_Data.xlsx", data_only=True) #to read only values/ not formulae
wbo = lw(filename = "RegDB_Out_Data.xlsx")
#print wbi.get_sheet_names()

#REGULATORY DATABASE VERSION 
ver = wbi.get_sheet_by_name('REGDB_VERSION_HISTORY')
versionRegDb = ver.cell(row=1,column=3).value

#REGULATORY DOMAINS 2G LOOKUP
ws = wbi.get_sheet_by_name('REG_DOMAINS_2G_LOOKUP')
REG_DOMAINS_2G_LOOKUP = {}
for i in range(2,ws.max_row+1):
    REG_DOMAINS_2G_LOOKUP[ws["A"+str(i)].value] = ws["B"+str(i)].value

#REGULATORY DOMAINS 5G LOOKUP
ws = wbi.get_sheet_by_name('REG_DOMAINS_5G_LOOKUP')
REG_DOMAINS_5G_LOOKUP = {}
for i in range(2,ws.max_row+1):
    REG_DOMAINS_5G_LOOKUP[ws["A"+str(i)].value] = ws["B"+str(i)].value
    
#REGULATORY RULES 2G LOOKUP
ws = wbi.get_sheet_by_name('REG_RULES_2G_LOOKUP')
REG_RULES_2G_LOOKUP = {}
for i in range(2,ws.max_row+1):
    REG_RULES_2G_LOOKUP[ws["A"+str(i)].value] = ws["B"+str(i)].value

#REGULATORY RULES 5G LOOKUP
ws = wbi.get_sheet_by_name('REG_RULES_5G_LOOKUP')
REG_RULES_5G_LOOKUP = {}
for i in range(2,ws.max_row+1):
    REG_RULES_5G_LOOKUP[ws["A"+str(i)].value] = ws["B"+str(i)].value

#IMPORTANT MACROS LOOKUP
ws = wbi.get_sheet_by_name('IMPORTANT_MACROS_LOOKUP')
IMPORTANT_MACROS_LOOKUP = {}
for i in range(2,ws.max_row+1):
    IMPORTANT_MACROS_LOOKUP[ws["A"+str(i)].value] = ws["C"+str(i)].value

#DFS LOOKUP
ws = wbi.get_sheet_by_name('DFS_LOOKUP')
DFS_LOOKUP = {}
for i in range(2,ws.max_row+1):
    DFS_LOOKUP[ws["A"+str(i)].value] = ws["B"+str(i)].value

#CTL LOOKUP
ws = wbi.get_sheet_by_name('CTL_LOOKUP')
CTL_LOOKUP = {}
for i in range(2,ws.max_row+1):
    CTL_LOOKUP[ws["A"+str(i)].value] = ws["C"+str(i)].value

#COUNTRY CODE LOOKUP
ws = wbi.get_sheet_by_name('COUNTRY_LOOKUP')
COUNTRY_LOOKUP = {}
for i in range(2,ws.max_row+1):
    COUNTRY_LOOKUP[ws["A"+str(i)].value] = ws["B"+str(i)].value

#REGULATORY DOMAIN PAIR LOOKUP
ws = wbi.get_sheet_by_name('REG_DMN_PAIR_LOOKUP')
REG_DMN_PAIR_LOOKUP = {}
for i in range(2,ws.max_row+1):
    REG_DMN_PAIR_LOOKUP[ws["A"+str(i)].value] = ws["C"+str(i)].value
#    REG_DMN_PAIR_LOOKUP[ws["A"+str(i)].value] = ws["C"+str(i)].internal_value # to get value instead of formulae - not working

# print REG_DOMAINS_2G_LOOKUP
# 
# print REG_RULES_2G_LOOKUP
# 
# print REG_DOMAINS_5G_LOOKUP
# 
# print REG_RULES_5G_LOOKUP
# 
# print IMPORTANT_MACROS_LOOKUP
# 
# print DFS_LOOKUP
#
# print CTL_LOOKUP
# 
# print COUNTRY_LOOKUP
# 
# print REG_DMN_PAIR_LOOKUP


# GLOBAL COUNTRIES DATABASE CONVERSION TO BDF FLATTENED FORMAT

wi = wbi.get_sheet_by_name('ALL_COUNTRIES_INPUT')
wo = wbo.get_sheet_by_name('ALL_COUNTRIES_OUTPUT')

alpha_len = 4 #additional alpha len added to flattened rows. only for country lookup

put_zeroes_column_in_sheet(wo, 'B')

for i in range(1,wi.max_row):
#    print "loop "+str(i)
    wo.cell(row=(1+((i-1)*(wi.max_column+alpha_len))+(1)),column=2).value=COUNTRY_LOOKUP[wi['A'+str(i+1)].value]
    #print Lookup1[wi['A'+str(i)].value]
    wo.cell(row=(1+((i-1)*(wi.max_column+alpha_len))+(2)),column=2).value=REG_DMN_PAIR_LOOKUP[wi['B'+str(i+1)].value]
    #print Lookup2[wi['B'+str(i)].value]
    charVal = wi['C'+str(i+1)].value
    finVal = ""
    if len(str(charVal)) == 3:
        for k in range(0,3):
            finVal = str('{:x}'.format(ord(charVal[k])))
            finVal = int("0x"+finVal,16)
            wo.cell(row=(1+((i-1)*(wi.max_column+alpha_len))+(3+k)),column=2).value=finVal
    else:
        for k in range(0,2):
            finVal = str('{:x}'.format(ord(charVal[k])))
            finVal = int("0x"+finVal,16)
            wo.cell(row=(1+((i-1)*(wi.max_column+alpha_len))+(3+k)),column=2).value=finVal
        wo.cell(row=(1+((i-1)*(wi.max_column+alpha_len))+(5)),column=2).value=0
#    print finVal
    
    charVal = wi['D'+str(i+1)].value
    finVal = ""
    if len(str(charVal)) == 3:
        for k in range(0,3):
            finVal = str('{:x}'.format(ord(charVal[k])))
            finVal = int("0x"+finVal,16)
            wo.cell(row=(1+((i-1)*(wi.max_column))+(6+k)),column=2).value=finVal
    else:
        for k in range(0,2):
            finVal = str('{:x}'.format(ord(charVal[k])))
            finVal = int("0x"+finVal,16)
            wo.cell(row=(1+((i-1)*(wi.max_column+alpha_len))+(6+k)),column=2).value=finVal
        wo.cell(row=(1+((i-1)*(wi.max_column+alpha_len))+(8)),column=2).value=0
#    print finVal
    
    wo.cell(row=(1+((i-1)*(wi.max_column+alpha_len))+(9)),column=2).value=wi['E'+str(i+1)].value
    
    wo.cell(row=(1+((i-1)*(wi.max_column+alpha_len))+(10)),column=2).value=wi['F'+str(i+1)].value
    
    if is_number(str(wi['G'+str(i+1)].value)):
        wo.cell(row=(1+((i-1)*(wi.max_column+alpha_len))+(11)),column=2).value=wi['G'+str(i+1)].value
    else:
        wo.cell(row=(1+((i-1)*(wi.max_column+alpha_len))+(12)),column=2).value=IMPORTANT_MACROS_LOOKUP[wi['G'+str(i+1)].value]
    
    wo.cell(row=(1+((i-1)*(wi.max_column+alpha_len))+(13)),column=2).value=wi['H'+str(i+1)].value
    
# REGULATORY DOMAIN PAIRS CONVERSION TO BDF FLATTENED FORMAT

wi = wbi.get_sheet_by_name('REG_DOMAIN_PAIRS_INPUT')
wo = wbo.get_sheet_by_name('REG_DOMAIN_PAIRS_OUTPUT')

put_zeroes_column_in_sheet(wo, 'B')

for i in range(1,wi.max_row):
#    print "loop "+str(i)
    wo.cell(row=(1+((i-1)*(wi.max_column))+(1)),column=2).value=REG_DMN_PAIR_LOOKUP[wi['A'+str(i+1)].value]
    wo.cell(row=(1+((i-1)*(wi.max_column))+(2)),column=2).value=REG_DOMAINS_5G_LOOKUP[wi['B'+str(i+1)].value]
    wo.cell(row=(1+((i-1)*(wi.max_column))+(3)),column=2).value=REG_DOMAINS_2G_LOOKUP[wi['C'+str(i+1)].value]
    
# REGULATORY RULES 2G CONVERSION TO BDF FLATTENED FORMAT

wi = wbi.get_sheet_by_name('REG_RULES_2G_INPUT')
wo = wbo.get_sheet_by_name('REG_RULES_2G_OUTPUT')

put_zeroes_column_in_sheet(wo, 'B')

for i in range(1,wi.max_row):
#    print "loop "+str(i)
#     wo.cell(row=(1+((i-1)*(wi.max_column))+(1)),column=2).value=wi['A'+str(i+1)].value
#     wo.cell(row=(1+((i-1)*(wi.max_column))+(2)),column=2).value=wi['B'+str(i+1)].value
#     wo.cell(row=(1+((i-1)*(wi.max_column))+(3)),column=2).value=wi['C'+str(i+1)].value
#     wo.cell(row=(1+((i-1)*(wi.max_column))+(4)),column=2).value=wi['D'+str(i+1)].value
#     wo.cell(row=(1+((i-1)*(wi.max_column))+(5)),column=2).value=wi['E'+str(i+1)].value
    for j in range(1,6):
        wo.cell(row=(1+((i-1)*(wi.max_column))+(j)),column=2).value=wi.cell(row=int(str(i+1)), column=j).value
     
    if is_number(str(wi['F'+str(i+1)].value)):
        wo.cell(row=(1+((i-1)*(wi.max_column))+(6)),column=2).value=wi['F'+str(i+1)].value
    else:
        wo.cell(row=(1+((i-1)*(wi.max_column))+(6)),column=2).value=IMPORTANT_MACROS_LOOKUP[wi['F'+str(i+1)].value]
 

# REGULATORY RULES 5G CONVERSION TO BDF FLATTENED FORMAT

wi = wbi.get_sheet_by_name('REG_RULES_5G_INPUT')
wo = wbo.get_sheet_by_name('REG_RULES_5G_OUTPUT')

put_zeroes_column_in_sheet(wo, 'B')

for i in range(1,wi.max_row):
#    print "loop "+str(i)
    for j in range(1,6):
        wo.cell(row=(1+((i-1)*(wi.max_column))+(j)),column=2).value=wi.cell(row=int(str(i+1)), column=j).value
        
    if is_number(str(wi['F'+str(i+1)].value)):
        wo.cell(row=(1+((i-1)*(wi.max_column))+(6)),column=2).value=wi['F'+str(i+1)].value
    else:
        wo.cell(row=(1+((i-1)*(wi.max_column))+(6)),column=2).value=IMPORTANT_MACROS_LOOKUP[wi['F'+str(i+1)].value]


# REGULATORY DOMAINS 2G CONVERSION TO BDF FLATTENED FORMAT

wi = wbi.get_sheet_by_name('REG_DOMAINS_2G_INPUT')
wo = wbo.get_sheet_by_name('REG_DOMAINS_2G_OUTPUT')

put_zeroes_column_in_sheet(wo, 'B')

for i in range(1,wi.max_row):
#    print "loop "+str(i)
    if is_number(str(wi['A'+str(i+1)].value)):
        wo.cell(row=(1+((i-1)*(wi.max_column))+(1)),column=2).value=wi['A'+str(i+1)].value
    else:
        wo.cell(row=(1+((i-1)*(wi.max_column))+(1)),column=2).value=CTL_LOOKUP[wi['A'+str(i+1)].value]
        
    if is_number(str(wi['B'+str(i+1)].value)):
        wo.cell(row=(1+((i-1)*(wi.max_column))+(2)),column=2).value=wi['B'+str(i+1)].value
    else:
        wo.cell(row=(1+((i-1)*(wi.max_column))+(2)),column=2).value=DFS_LOOKUP[wi['B'+str(i+1)].value]
        
    for j in range(3,7):
        wo.cell(row=(1+((i-1)*(wi.max_column))+(j)),column=2).value=wi.cell(row=int(str(i+1)), column=j).value
    
    for j in range(7,18):
        if wi.cell(row=int(str(i+1)), column=j).value == None:
            break
        else:    
            if is_number(str(wi.cell(row=int(str(i+1)), column=j).value)):
                wo.cell(row=(1+((i-1)*(wi.max_column))+(j)),column=2).value=wi.cell(row=int(str(i+1)), column=j).value
            else:
                wo.cell(row=(1+((i-1)*(wi.max_column))+(j)),column=2).value=REG_RULES_2G_LOOKUP[wi.cell(row=int(str(i+1)), column=j).value]
         
# REGULATORY DOMAINS 2G CONVERSION TO BDF FLATTENED FORMAT

wi = wbi.get_sheet_by_name('REG_DOMAINS_5G_INPUT')
wo = wbo.get_sheet_by_name('REG_DOMAINS_5G_OUTPUT')

put_zeroes_column_in_sheet(wo, 'B')

for i in range(1,wi.max_row):
#    print "loop "+str(i)
    if is_number(str(wi['A'+str(i+1)].value)):
        wo.cell(row=(1+((i-1)*(wi.max_column))+(1)),column=2).value=wi['A'+str(i+1)].value
    else:
        wo.cell(row=(1+((i-1)*(wi.max_column))+(1)),column=2).value=CTL_LOOKUP[wi['A'+str(i+1)].value]
        
    if is_number(str(wi['B'+str(i+1)].value)):
        wo.cell(row=(1+((i-1)*(wi.max_column))+(2)),column=2).value=wi['B'+str(i+1)].value
    else:
        wo.cell(row=(1+((i-1)*(wi.max_column))+(2)),column=2).value=DFS_LOOKUP[wi['B'+str(i+1)].value]
        
    for j in range(3,7):
        wo.cell(row=(1+((i-1)*(wi.max_column))+(j)),column=2).value=wi.cell(row=int(str(i+1)), column=j).value
    
    for j in range(7,18):
        if wi.cell(row=int(str(i+1)), column=j).value == None:
            break
        else:    
            if is_number(str(wi.cell(row=int(str(i+1)), column=j).value)):
                wo.cell(row=(1+((i-1)*(wi.max_column))+(j)),column=2).value=wi.cell(row=int(str(i+1)), column=j).value
            else:
                wo.cell(row=(1+((i-1)*(wi.max_column))+(j)),column=2).value=REG_RULES_5G_LOOKUP[wi.cell(row=int(str(i+1)), column=j).value]

#PUTTING ALL SHEETS OUTPUT IN ONE SHEET

wo = wbo.get_sheet_by_name('COMBINED_REGULATORY_FIELDS') 

put_zeroes_column_in_sheet(wo, 'B')

combine_sheets('ALL_COUNTRIES_OUTPUT','COMBINED_REGULATORY_FIELDS',wbo,"regDbAllCountries")
combine_sheets('REG_DOMAIN_PAIRS_OUTPUT','COMBINED_REGULATORY_FIELDS',wbo,"regDbRegDmnPairs")
combine_sheets('REG_RULES_2G_OUTPUT','COMBINED_REGULATORY_FIELDS',wbo,"regDbRegRule2g")
combine_sheets('REG_RULES_5G_OUTPUT','COMBINED_REGULATORY_FIELDS',wbo,"regDbRegRule5g")
combine_sheets('REG_DOMAINS_2G_OUTPUT','COMBINED_REGULATORY_FIELDS',wbo,"regDbRegDomains2g")
combine_sheets('REG_DOMAINS_5G_OUTPUT','COMBINED_REGULATORY_FIELDS',wbo,"regDbRegDomains5g")

#ID & Len for reg DB section

wo.cell(row=2,column=2).value = int(20)
wo.cell(row=3,column=2).value = int(7172)


print "done"    
wbo.save("RegDB_Out_Data.xlsx")
wbo.close()
wbi.close()

#pushing values to TXT file
wbi = lw(filename = "RegDB_Out_Data.xlsx") #to read only values/ not formulae
#wbo = lw(filename = NvData, read_only=False, keep_vba=True)

txtLines = list()
with open("regdb.txt", "r") as txt:
    txtLines = txt.readlines()


wi = wbi.get_sheet_by_name("COMBINED_REGULATORY_FIELDS")   
#wo = wbo.get_sheet_by_name("Data") 

for i in range(len(txtLines)):
    if (txtLines[i].find(wi.cell(row=5,column=1).value) != -1):
        break

for j in range(4,wi.max_row): #starts from 4th row, to balance unused values in output excel file
    l = txtLines[i+j-3-1].split("\t")
    l[2] = l[2].strip()
    l[2] = str(wi.cell(row=j+1,column=2).value)+"\n" 
    txtLines[i+j-3-1] = "\t".join(l)

####RegDb is enabled by default####
l = txtLines[i+j-3-1].split("\t") #Always enable flag is expected immediately after database ends
l[2] = l[2].strip()
l[2] = "1"+"\n"
txtLines[i+j-3-1] = "\t".join(l)

####RegDb version setting####
l = txtLines[i+j-3].split("\t") #Always version field is expected after enable/disable flag
l[2] = l[2].strip()
######Setting value######
l[2] = str(versionRegDb)+"\n"
txtLines[i+j-3] = "\t".join(l)

with open("regdb.txt","w") as txt:
    txt.truncate
    for k in txtLines:
        txt.writelines(k)
        #txt.writelines(" \n")


#wbo.save(NvData)
#wbo.close()
wbi.close()

try:
    subprocess.call("python RegDB_txt2bin.py regdb.txt regdb.bin")
except:
    print ("Error converting txt to bin")
    exit()
try:
    subprocess.call("python RegDB_bin2txt.py regdb.bin regdb.txt")
except:
    print ("Error converting bin to txt")
    exit()