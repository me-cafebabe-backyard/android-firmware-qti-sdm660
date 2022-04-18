#import timeit # very important to calculate time complexities
from sys import argv
import sys
import os
import subprocess
import fileinput
import re
import time
import argparse
from openpyxl import load_workbook as lw
from openpyxl import Workbook 
from openpyxl import load_workbook
from openpyxl.styles import Font

#### NEED TO FIX THESE IN NEXT RELEASE ####
'''print "\nWarning : Instead of REGULATORY_CHAN_NO11N , pls modify it to REGULATORY_PHYMODE_NO11N in input excel file"
print "Issue1 : Repeated countries Ex: CTRY_LIBYA, 434, twice"
print "Warning : Need to use 3 dictionaries each for CHAN, PHYMODE and combination of both?\n"'''

#---------------------------------------------
# This following logic suppresses the warnings 
def warn(*args, **kwargs):
    pass
import warnings
warnings.warn = warn
#---------------------------------------------

d = os.path.dirname(os.getcwd())
d1 = os.path.dirname(d)
split = 0
regDbTxtLines = []

#Prompt user to specify whether it is split BDF or combined BDF
parser = argparse.ArgumentParser(description='regDb Parser from bdf')
parser.add_argument('inputExcelInfile', metavar='<Regulatory_BDF_In_Data>.xlsx', help='Mandatory : provide DataIn Excel file containing regDb')
parser.add_argument('inputfile', metavar='<bdwlan>.txt', help='Mandatory : provide the BDF file containing regDb')
#group = parser.add_mutually_exclusive_group()
#group.add_argument('-v', '--verbose', action='store_true', help='print verbose')
args = parser.parse_args()

ext1 = args.inputExcelInfile.lower().rsplit('.', 1)[-1]
ext2 = args.inputfile.lower().rsplit('.', 1)[-1]
if (ext1 != "xlsx" or ext2 != "txt"):
	print "\nInput Argument Error : Please check usage and provide valid arguments in that order and re run the tool"
	exit()

if not os.path.isfile(args.inputExcelInfile):
	print "\nInput Error : Input Data In Excel file " +str(args.inputExcelInfile)+ " does not exist"
	exit()

if not os.path.isfile(args.inputfile):
	print "\nInput Error : Input BDF file " + str(args.inputfile) + " does not exist"
	exit()

split = raw_input("\nPlease make following selection: \n\n 1. Input file is Split BDF\n 2. Input file is Complete BDF \n 3. Exit the Tool\n\nSelect your option : ")

#Some Important Global Definitions here (These macros need to be updated as per those in common_bdf.h)
#--------------------------------------
NUM_REG_COUNTRIES 		= 0
NUM_REG_DOMAIN_PAIRS	= 0
NUM_REG_RULES_2G		= 0
NUM_REG_RULES_5G		= 0
NUM_REG_DOMAINS_2G		= 0
NUM_REG_DOMAINS_5G		= 0

MAX_REG_RULES  = 10	# Current Max Rules is set to 10, but it can be changed, 
					# so when the input excel gets changed, modify this field accordingly
REG_ALPHA2_LEN = 2



#Read the Input Text File for Regulatory
#1.Input Split BDF Text file : Regulatory ONLY
if (split == '1') :
	with open(args.inputfile, "r") as inFile:
		regDbTxtLines_comb = [line.split() for line in inFile]

	del regDbTxtLines_comb[0]

	for line in regDbTxtLines_comb:
		if(line[0] != "uint16"):	# split BDF 
			print "\nInput Error : The Input file is not a split BDF, please run the tool again with a valid split BDF"
			exit(0)
		else:
			break

	z=0
	for line in regDbTxtLines_comb:
		if(line[1] == "REGULATORY_DB_SECTION.regulatorytDbFuture[0]"):
			break
		regDbTxtLines.append(line)
		z += 1

	NUM_REG_COUNTRIES 		= 660
	NUM_REG_DOMAIN_PAIRS	= 450
	NUM_REG_RULES_2G		= 45
	NUM_REG_RULES_5G		= 270
	NUM_REG_DOMAINS_2G		= 45
	NUM_REG_DOMAINS_5G		= 210

#2.Input BDF file with Regulatory
elif(split == '2'):
	with open(args.inputfile, "r") as inFile:
		regDbTxtLines_comb = [line.split() for line in inFile]

	for line in regDbTxtLines_comb:
		if(line[0] != "nvBaseId__0_0"):	# split BDF 
			print "\nInput Error : The Input file is not a complete BDF, please run the tool again with a valid complete BDF"
			exit(0)
		else:
			break

	# Search for the first field of regulatory section
        regDbStart = False
        startDet = True
        endDet = False
        for line in regDbTxtLines_comb:
            if startDet and (str(line[0]) == 'REGULATORY_DB_SECTION.regDbAllCountries.country_code__0_0'):
                regDbStart = True
                startDet = False
                endDet = True

            if endDet and (str(line[0]) == 'regulatorytDbFuture'):
            	regDbStart = False
            	break                                    

            if(regDbStart):
            	regDbTxtLines.append(line)

    	NUM_REG_COUNTRIES 		= 220
	NUM_REG_DOMAIN_PAIRS	= 150
	NUM_REG_RULES_2G		= 15
	NUM_REG_RULES_5G		= 90
	NUM_REG_DOMAINS_2G		= 15
	NUM_REG_DOMAINS_5G		= 70       

else:
	print "Invalid selection, please make valid selection"
	exit(0)            	

#3.Remove the RegDbFlag and regDbVersion from regDbTxtLines list
str1 = regDbTxtLines[-1]
if(split == '1'):
	regDbVersion = hex(int(str(str1[2])))
else:
	regDbVersion = str(str1[1])
del regDbTxtLines[-1]

str1 = regDbTxtLines[-1]
if(split == '1'):
	regDbFlag = str(str1[2])
else:
	regDbFlag = str(str1[1])
del regDbTxtLines[-1]


#4.Strip each line to separate country codes, reg domain pairs etc.
regDbAllCountriesValues = []
regDbRegDmnPairsValues = []
regDbRegRuleValues = []
regDbRegDomainsValues = [] 

idx = 0
for line in regDbTxtLines:	
	if(split == '1'):
		string = str(line[1]).split('.')
		key = string[1].split('[')
		key = key[0]                
		value = line[2]
	else:
	    string = str(line[0]).split('.')
	    key = string[1]
	    value = line[1]
	
        if (key == 'regDbAllCountries'):
            regDbAllCountriesValues.append(str(value))
        elif (key == 'regDbRegDmnPairs'):
            regDbRegDmnPairsValues.append(str(value))
        elif (key == 'regDbRegRule' or key == 'regDbRegRule2g' or key == 'regDbRegRule5g'):
            regDbRegRuleValues.append(str(value))
        elif (key == 'regDbRegDomains' or key == 'regDbRegDomains2g' or key == 'regDbRegDomains5g'):
            regDbRegDomainsValues.append(str(value))

#5.Modify the "Regulatory_BDF_In_Data.xlsx" with regulatory values from input BDF/Split BDF
wbo = Workbook()
filepath = "Regulatory_BDF_Parsed_Data.xlsx"
wbi = lw(filename = args.inputExcelInfile, data_only=True)

#5.1 REGDB_VERSION_HISTORY
veri = wbi.get_sheet_by_name('REGDB_VERSION_HISTORY')
vero = wbo.active
vero.title = 'REGDB_VERSION_HISTORY'
vero.cell(row=1, column=1).value = veri.cell(row=1, column=1).value
vero.cell(row=1, column=2).value = veri.cell(row=1, column=2).value
vero.cell(row=1, column=3).value = veri.cell(row=1, column=3).value
vero.cell(row=1, column=4).value = veri.cell(row=1, column=4).value

#5.2 REG_DOMAINS_2G_LOOKUP
regDom2GLookUp = {}
regD2g = wbi.get_sheet_by_name('REG_DOMAINS_2G_LOOKUP')
regD2g_out = wbo.create_sheet('REG_DOMAINS_2G_LOOKUP')

regD2g_out.cell(row=1, column=1).value = regD2g.cell(row=1, column=1).value	
regD2g_out.cell(row=1, column=2).value = regD2g.cell(row=1, column=2).value
for i in range(2, regD2g.max_row+1):	 
	key = regD2g.cell(row=i, column=1).value
	value = regD2g.cell(row=i, column=2).value
	regD2g_out.cell(row=i, column=1).value = key	
	regD2g_out.cell(row=i, column=2).value = value
	regDom2GLookUp[value] = key

#5.3 REG_DOMAINS_5G_LOOKUP
regDom5GLookUp = {}
regD5g = wbi.get_sheet_by_name('REG_DOMAINS_5G_LOOKUP')
regD5g_out = wbo.create_sheet('REG_DOMAINS_5G_LOOKUP')

regD5g_out.cell(row=1, column=1).value = regD5g.cell(row=1, column=1).value
regD5g_out.cell(row=1, column=2).value = regD5g.cell(row=1, column=2).value
for i in range(2, regD5g.max_row+1):
	key = regD5g.cell(row=i, column=1).value
	value = regD5g.cell(row=i, column=2).value
	regD5g_out.cell(row=i, column=1).value = key
	regD5g_out.cell(row=i, column=2).value = value 
	regDom5GLookUp[value] = key

#5.4 REG_RULES_2G_LOOKUP
regRul2GLookUp = {}
regR2g = wbi.get_sheet_by_name('REG_RULES_2G_LOOKUP')
regR2g_out = wbo.create_sheet('REG_RULES_2G_LOOKUP')

regR2g_out.cell(row=1, column=1).value = regR2g.cell(row=1, column=1).value
regR2g_out.cell(row=1, column=2).value = regR2g.cell(row=1, column=2).value
for i in range(2, regR2g.max_row+1):
	key = regR2g.cell(row=i, column=1).value
	value = regR2g.cell(row=i, column=2).value
	regR2g_out.cell(row=i, column=1).value = key
	regR2g_out.cell(row=i, column=2).value = value 
	regRul2GLookUp[value] = key

#5.5 REG_RULES_5G_LOOKUP
regRul5GLookUp = {}
regR5g = wbi.get_sheet_by_name('REG_RULES_5G_LOOKUP')
regR5g_out = wbo.create_sheet('REG_RULES_5G_LOOKUP')

regR5g_out.cell(row=1, column=1).value = regR5g.cell(row=1, column=1).value
regR5g_out.cell(row=1, column=2).value = regR5g.cell(row=1, column=2).value
for i in range(2, regR5g.max_row+1):
	key = regR5g.cell(row=i, column=1).value
	value = regR5g.cell(row=i, column=2).value
	regR5g_out.cell(row=i, column=1).value = key
	regR5g_out.cell(row=i, column=2).value = value 
	regRul5GLookUp[value] = key

#5.6 IMPORTANT_MACROS_LOOKUP
macrosLookUp_Chan = {}
macrosLookUp_PhyMode = {}
macros = wbi.get_sheet_by_name('IMPORTANT_MACROS_LOOKUP')
macros_out = wbo.create_sheet('IMPORTANT_MACROS_LOOKUP')

macros_out.cell(row=1, column=1).value = macros.cell(row=1, column=1).value
macros_out.cell(row=1, column=2).value = macros.cell(row=1, column=2).value
macros_out.cell(row=1, column=3).value = macros.cell(row=1, column=3).value
for i in range(2, macros.max_row+1):
	key = macros.cell(row=i, column=1).value
	value = macros.cell(row=i, column=3).value
	macros_out.cell(row=i, column=1).value = key
	macros_out.cell(row=i, column=2).value = macros.cell(row=i, column=2).value
	macros_out.cell(row=i, column=3).value = value
	string = key.split('_')
	if(string[1] == "CHAN"):
	    macrosLookUp_Chan[value] = key
	elif(string[1] == "PHYMODE"):
	    macrosLookUp_PhyMode[value] = key
	else:
            print "Unknown Macro"
            exit(0)

#5.7 DFS_LOOKUP
DfsLookUp = {}
dfs = wbi.get_sheet_by_name('DFS_LOOKUP')
dfs_out = wbo.create_sheet('DFS_LOOKUP')

dfs_out.cell(row=1, column=1).value = dfs.cell(row=1, column=1).value
dfs_out.cell(row=1, column=2).value = dfs.cell(row=1, column=2).value
for i in range(2, dfs.max_row+1):
	key = dfs.cell(row=i, column=1).value
	value = dfs.cell(row=i, column=2).value
	dfs_out.cell(row=i, column=1).value = key
	dfs_out.cell(row=i, column=2).value = value
	DfsLookUp[value] = key

#5.8 CTL_LOOKUP
CtlLookUp = {}
ctl = wbi.get_sheet_by_name('CTL_LOOKUP')
ctl_out = wbo.create_sheet('CTL_LOOKUP')

ctl_out.cell(row=1, column=1).value = ctl.cell(row=1, column=1).value
ctl_out.cell(row=1, column=2).value = ctl.cell(row=1, column=2).value
ctl_out.cell(row=1, column=3).value = ctl.cell(row=1, column=3).value
for i in range(2, ctl.max_row+1):
	key = ctl.cell(row=i, column=1).value
	value = ctl.cell(row=i, column=3).value
	ctl_out.cell(row=i, column=1).value = key
	ctl_out.cell(row=i, column=2).value = ctl.cell(row=i, column=2).value
	ctl_out.cell(row=i, column=3).value = value
	CtlLookUp[value] = key

#5.9 COUNTRY_LOOKUP
CntryLookUp = {}
cntry = wbi.get_sheet_by_name('COUNTRY_LOOKUP')
cntry_out = wbo.create_sheet('COUNTRY_LOOKUP')

cntry_out.cell(row=1, column=1).value = cntry.cell(row=1, column=1).value
cntry_out.cell(row=1, column=2).value = cntry.cell(row=1, column=2).value
for i in range(2, cntry.max_row+1):
	key = cntry.cell(row=i, column=1).value
	value = cntry.cell(row=i, column=2).value
	cntry_out.cell(row=i, column=1).value = key
	cntry_out.cell(row=i, column=2).value = value
	CntryLookUp[value] = key 

#5.10 REG_DMN_PAIR_LOOKUP
RegDmnLookUp = {}
regDmn = wbi.get_sheet_by_name('REG_DMN_PAIR_LOOKUP')
regDmn_out = wbo.create_sheet('REG_DMN_PAIR_LOOKUP')

regDmn_out.cell(row=1, column=1).value = regDmn.cell(row=1, column=1).value
regDmn_out.cell(row=1, column=2).value = regDmn.cell(row=1, column=2).value
regDmn_out.cell(row=1, column=3).value = regDmn.cell(row=1, column=3).value
for i in range(1, regDmn.max_row+1):
	key = regDmn.cell(row=i, column=1).value
	value = regDmn.cell(row=i, column=3).value
	regDmn_out.cell(row=i, column=1).value = key
	regDmn_out.cell(row=i, column=2).value = regDmn.cell(row=i, column=2).value
	regDmn_out.cell(row=i, column=3).value = value 
	RegDmnLookUp[value] = key

#5.12 ALL_COUNTRIES_INPUT
def all_countries_input(values, idx, curr_out_sheet):
	if(int(values[idx*12]) > 0):
		temp1 = int(values[idx*12+0])
		if(temp1 in CntryLookUp):
			curr_out_sheet.cell(row=idx+2, column=1).value = CntryLookUp[temp1]
		else:
			curr_out_sheet.cell(row=idx+2, column=1).value = str("UNDEFINED ")+str("[")+str(temp1)+str("]")

		temp2 = int(values[idx*12+1])
		if(temp2 in RegDmnLookUp):
			curr_out_sheet.cell(row=idx+2, column=2).value = RegDmnLookUp[temp2]
		else:
			curr_out_sheet.cell(row=idx+2, column=2).value = str("UNDEFINED ")+str("[")+str(temp2)+str("]")

		ascii1 = int(values[idx*12+2])
		ascii2 = int(values[idx*12+3])
		ascii3 = int(values[idx*12+4])
		if(ascii1 > 0): 
			alpha2_1 = chr(ascii1) 	#Get the character given by an ASCII number
		else: 
			alpha2_1 = ""
		if(ascii2 > 0): 
			alpha2_2 = chr(ascii2)
		else: 
			alpha2_2 = ""
		if(ascii3 > 0): 
			alpha2_3 = chr(ascii3)
		else: 
			alpha2_3 = ""

		curr_out_sheet.cell(row=idx+2, column=3).value = alpha2_1 + alpha2_2 + alpha2_3

		ascii1 = int(values[idx*12+5])
		ascii2 = int(values[idx*12+6])
		ascii3 = int(values[idx*12+7])
		if(ascii1 > 0): 
			alpha2_11D_1 = chr(ascii1) 
		else: 
			alpha2_11D_1 = ""
		if(ascii2 > 0): 
			alpha2_11D_2 = chr(ascii2) 
		else: 
			alpha2_11D_2 = ""
		if(ascii3 > 0): 
			alpha2_11D_3 = chr(ascii3)
		else: 
			alpha2_11D_3 = ""

		curr_out_sheet.cell(row=idx+2, column=4).value = alpha2_11D_1 + alpha2_11D_2 + alpha2_11D_3

		curr_out_sheet.cell(row=idx+2, column=5).value = int(values[idx*12+8])
		curr_out_sheet.cell(row=idx+2, column=6).value = int(values[idx*12+9])
		curr_out_sheet.cell(row=idx+2, column=7).value = int(values[idx*12+10])
		curr_out_sheet.cell(row=idx+2, column=8).value = int(values[idx*12+11])

allCntry = wbi.get_sheet_by_name('ALL_COUNTRIES_INPUT')
allCntry_out = wbo.create_sheet('ALL_COUNTRIES_INPUT')
allCntry_out.cell(row=1, column=1).value = allCntry.cell(row=1, column=1).value
allCntry_out.cell(row=1, column=2).value = allCntry.cell(row=1, column=2).value
allCntry_out.cell(row=1, column=3).value = allCntry.cell(row=1, column=3).value
allCntry_out.cell(row=1, column=4).value = allCntry.cell(row=1, column=4).value
allCntry_out.cell(row=1, column=5).value = allCntry.cell(row=1, column=5).value
allCntry_out.cell(row=1, column=6).value = allCntry.cell(row=1, column=6).value
allCntry_out.cell(row=1, column=7).value = allCntry.cell(row=1, column=7).value
allCntry_out.cell(row=1, column=8).value = allCntry.cell(row=1, column=8).value

for k in range(0, NUM_REG_COUNTRIES):
	all_countries_input(regDbAllCountriesValues, k, allCntry_out)	

#5.13 REG_DOMAIN_PAIRS_INPUT
def reg_domain_pairs_input(values, idx, curr_out_sheet):
	if(int(values[idx*3+0]) > 0):
		temp1 = int(values[idx*3+0])
		if(temp1 in RegDmnLookUp):
			curr_out_sheet.cell(row=idx+2, column=1).value = RegDmnLookUp[temp1]
		else:
			curr_out_sheet.cell(row=idx+2, column=1).value = str("UNDEFINED ")+str("[")+str(temp1)+str("]")

		temp2 = int(values[idx*3+1])
		if(temp2 in regDom5GLookUp):
			curr_out_sheet.cell(row=idx+2, column=2).value = regDom5GLookUp[temp2]
		else:
			curr_out_sheet.cell(row=idx+2, column=2).value = str("UNDEFINED ")+str("[")+str(temp2)+str("]")

		temp3 = int(values[idx*3+2])
		if(temp3 in regDom2GLookUp):
			curr_out_sheet.cell(row=idx+2, column=3).value = regDom2GLookUp[temp3]
		else:
			curr_out_sheet.cell(row=idx+2, column=3).value = str("UNDEFINED ")+str("[")+str(temp3)+str("]")


regDmnPairs = wbi.get_sheet_by_name('REG_DOMAIN_PAIRS_INPUT')
regDmnPairs_out = wbo.create_sheet('REG_DOMAIN_PAIRS_INPUT')
regDmnPairs_out.cell(row=1, column=1).value = regDmnPairs.cell(row=1, column=1).value
regDmnPairs_out.cell(row=1, column=2).value = regDmnPairs.cell(row=1, column=2).value
regDmnPairs_out.cell(row=1, column=3).value = regDmnPairs.cell(row=1, column=3).value

for k in range(0, NUM_REG_DOMAIN_PAIRS):
	reg_domain_pairs_input(regDbRegDmnPairsValues, k, regDmnPairs_out)


#5.14 REG_RULES_2G_INPUT
def reg_rules_2g_input(values, idx, curr_out_sheet):
	if(int(values[idx*6+0]) > 0):
		curr_out_sheet.cell(row=idx+2, column=1).value = int(values[idx*6+0])
		curr_out_sheet.cell(row=idx+2, column=2).value = int(values[idx*6+1])
		curr_out_sheet.cell(row=idx+2, column=3).value = int(values[idx*6+2])
		curr_out_sheet.cell(row=idx+2, column=4).value = int(values[idx*6+3])
		curr_out_sheet.cell(row=idx+2, column=5).value = int(values[idx*6+4])

		temp1 = int(values[idx*6+5])
		if(temp1 > 0):
			if(temp1 in macrosLookUp_Chan):
				curr_out_sheet.cell(row=idx+2, column=6).value = macrosLookUp_Chan[temp1]
			else:
				curr_out_sheet.cell(row=idx+2, column=6).value = str("UNDEFINED ")+str("[")+str(temp1)+str("]")
		else:
			curr_out_sheet.cell(row=idx+2, column=6).value = 0

regRul2g = wbi.get_sheet_by_name('REG_RULES_2G_INPUT')
regRul2g_out = wbo.create_sheet('REG_RULES_2G_INPUT')
regRul2g_out.cell(row=1, column=1).value = regRul2g.cell(row=1, column=1).value
regRul2g_out.cell(row=1, column=2).value = regRul2g.cell(row=1, column=2).value
regRul2g_out.cell(row=1, column=3).value = regRul2g.cell(row=1, column=3).value
regRul2g_out.cell(row=1, column=4).value = regRul2g.cell(row=1, column=4).value
regRul2g_out.cell(row=1, column=5).value = regRul2g.cell(row=1, column=5).value
regRul2g_out.cell(row=1, column=6).value = regRul2g.cell(row=1, column=6).value

for k in range(0, NUM_REG_RULES_2G):
	reg_rules_2g_input(regDbRegRuleValues, k, regRul2g_out)

#5.15 REG_DOMAINS_2G_INPUT
def reg_domains_2g_input(values, idx, curr_out_sheet):
	if(int(values[idx*16+0]) > 0):
		temp1 = int(values[idx*16+0])
		if(temp1 in CtlLookUp):
			curr_out_sheet.cell(row=idx+2, column=1).value = CtlLookUp[temp1]
		else:
			curr_out_sheet.cell(row=idx+2, column=1).value = str("UNDEFINED ")+str("[")+str(temp1)+str("]")

		temp2 = int(values[idx*16+1])
		if(temp2 in DfsLookUp):
			curr_out_sheet.cell(row=idx+2, column=2).value = DfsLookUp[temp2]
		else:
			curr_out_sheet.cell(row=idx+2, column=2).value = str("UNDEFINED ")+str("[")+str(temp2)+str("]")

		curr_out_sheet.cell(row=idx+2, column=3).value = int(values[idx*16+2])
		curr_out_sheet.cell(row=idx+2, column=4).value = int(values[idx*16+3])
		curr_out_sheet.cell(row=idx+2, column=5).value = int(values[idx*16+4])
		curr_out_sheet.cell(row=idx+2, column=6).value = int(values[idx*16+5])

	for m in range(0, int(values[idx*16+4])):
		temp3 = int(values[idx*16+6+m])
		if(temp3 in regRul2GLookUp):
			curr_out_sheet.cell(row=idx+2, column=7+m).value = regRul2GLookUp[temp3]
		else:
			curr_out_sheet.cell(row=idx+2, column=7+m).value = str("UNDEFINED ")+str("[")+str(temp3)+str("]")


regDmn2g = wbi.get_sheet_by_name('REG_DOMAINS_2G_INPUT')
regDmn2g_out = wbo.create_sheet('REG_DOMAINS_2G_INPUT')
regDmn2g_out.cell(row=1, column=1).value = regDmn2g.cell(row=1, column=1).value
regDmn2g_out.cell(row=1, column=2).value = regDmn2g.cell(row=1, column=2).value
regDmn2g_out.cell(row=1, column=3).value = regDmn2g.cell(row=1, column=3).value
regDmn2g_out.cell(row=1, column=4).value = regDmn2g.cell(row=1, column=4).value
regDmn2g_out.cell(row=1, column=5).value = regDmn2g.cell(row=1, column=5).value
regDmn2g_out.cell(row=1, column=6).value = regDmn2g.cell(row=1, column=6).value
for i in range(0, MAX_REG_RULES):
	regDmn2g_out.cell(row=1, column=7+i).value = regDmn2g.cell(row=1, column=7+i).value

for k in range(0, NUM_REG_DOMAINS_2G):
	reg_domains_2g_input(regDbRegDomainsValues, k, regDmn2g_out)

#5.16 REG_RULES_5G_INPUT
def reg_rules_5g_input(values, idx, curr_out_sheet):
    diff = (idx*6) + (NUM_REG_RULES_2G*6)
    if(int(values[diff+0]) > 0):
		curr_out_sheet.cell(row=idx+2, column=1).value = int(values[diff+0])
		curr_out_sheet.cell(row=idx+2, column=2).value = int(values[diff+1])
		curr_out_sheet.cell(row=idx+2, column=3).value = int(values[diff+2])
		curr_out_sheet.cell(row=idx+2, column=4).value = int(values[diff+3])
		curr_out_sheet.cell(row=idx+2, column=5).value = int(values[diff+4])
		temp1 = int(values[diff+5])
		if(temp1 > 0):
			if(temp1 in macrosLookUp_Chan):
				curr_out_sheet.cell(row=idx+2, column=6).value = macrosLookUp_Chan[temp1]
			else:
				curr_out_sheet.cell(row=idx+2, column=6).value = str("UNDEFINED ")+str("[")+str(temp1)+str("]")
		else:
			curr_out_sheet.cell(row=idx+2, column=6).value = 0

regRul5g = wbi.get_sheet_by_name('REG_RULES_5G_INPUT')
regRul5g_out = wbo.create_sheet('REG_RULES_5G_INPUT')
regRul5g_out.cell(row=1, column=1).value = regRul5g.cell(row=1, column=1).value
regRul5g_out.cell(row=1, column=2).value = regRul5g.cell(row=1, column=2).value
regRul5g_out.cell(row=1, column=3).value = regRul5g.cell(row=1, column=3).value
regRul5g_out.cell(row=1, column=4).value = regRul5g.cell(row=1, column=4).value
regRul5g_out.cell(row=1, column=5).value = regRul5g.cell(row=1, column=5).value
regRul5g_out.cell(row=1, column=6).value = regRul5g.cell(row=1, column=6).value

for k in range(0, NUM_REG_RULES_5G):
	reg_rules_5g_input(regDbRegRuleValues, k, regRul5g_out)

#5.17 REG_DOMAINS_5G_INPUT
def reg_domains_5g_input(values, idx, curr_out_sheet):
	diff = (idx*16) + (NUM_REG_DOMAINS_2G*16)
	if(int(values[diff+0]) > 0):
		temp1 = int(values[diff+0])
		if(temp1 in CtlLookUp):
			curr_out_sheet.cell(row=idx+3, column=1).value = CtlLookUp[temp1]
		else:
			curr_out_sheet.cell(row=idx+3, column=1).value = str("UNDEFINED ")+str("[")+str(temp1)+str("]")

		temp2 = int(values[diff+1])
		if(temp2 in DfsLookUp):
			curr_out_sheet.cell(row=idx+3, column=2).value = DfsLookUp[temp2]
		else:
			curr_out_sheet.cell(row=idx+3, column=2).value = str("UNDEFINED ")+str("[")+str(temp2)+str("]")

		curr_out_sheet.cell(row=idx+3, column=3).value = int(values[diff+2])
		curr_out_sheet.cell(row=idx+3, column=4).value = int(values[diff+3])
		curr_out_sheet.cell(row=idx+3, column=5).value = int(values[diff+4])
		curr_out_sheet.cell(row=idx+3, column=6).value = int(values[diff+5])

	for m in range(0, int(values[diff+4])):
		temp3 = int(values[diff+6+m])
		if(temp3 in regRul5GLookUp):
			curr_out_sheet.cell(row=idx+3, column=7+m).value = regRul5GLookUp[temp3]
		else:
			curr_out_sheet.cell(row=idx+3, column=7+m).value = str("UNDEFINED ")+str("[")+str(temp3)+str("]")

regDmn5g = wbi.get_sheet_by_name('REG_DOMAINS_5G_INPUT')
regDmn5g_out = wbo.create_sheet('REG_DOMAINS_5G_INPUT')
regDmn5g_out.cell(row=1, column=1).value = regDmn5g.cell(row=1, column=1).value
regDmn5g_out.cell(row=2, column=1).value = 0
regDmn5g_out.cell(row=1, column=2).value = regDmn5g.cell(row=1, column=2).value
regDmn5g_out.cell(row=2, column=2).value = 0
regDmn5g_out.cell(row=1, column=3).value = regDmn5g.cell(row=1, column=3).value
regDmn5g_out.cell(row=2, column=3).value = 0
regDmn5g_out.cell(row=1, column=4).value = regDmn5g.cell(row=1, column=4).value
regDmn5g_out.cell(row=2, column=4).value = 0
regDmn5g_out.cell(row=1, column=5).value = regDmn5g.cell(row=1, column=5).value
regDmn5g_out.cell(row=2, column=5).value = 0
regDmn5g_out.cell(row=1, column=6).value = regDmn5g.cell(row=1, column=6).value
regDmn5g_out.cell(row=2, column=6).value = 0
for i in range(0, MAX_REG_RULES):
	regDmn5g_out.cell(row=1, column=7+i).value = regDmn5g.cell(row=1, column=7+i).value

regDmn5g_out.cell(row=2, column=7).value = 0

for k in range(0, NUM_REG_DOMAINS_5G):
	reg_domains_5g_input(regDbRegDomainsValues, k, regDmn5g_out)

wbo.save(filepath)

print "\nSuccessfully generated 'Regulatory_BDF_Parsed_Data.xlsx' file"
