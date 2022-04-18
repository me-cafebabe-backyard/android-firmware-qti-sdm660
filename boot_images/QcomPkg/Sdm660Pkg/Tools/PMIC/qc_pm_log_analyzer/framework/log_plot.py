#!/usr/bin/python
# ===== Py 2/3 Compatibility Layer =====
from __future__ import absolute_import, division, print_function
from future.builtins import (ascii, bytes, chr, dict, filter, hex, input,
                             int, map, next, oct, open, pow, range, round,
                             str, super, zip)
from future.builtins.disabled import *
# ======================================
from framework.log_analyzer_config import LogAnalyzerConfig, PlotDataSeries
from framework.log_data_set import LogGroupSet
from framework.log_parser import LogParser
from openpyxl import Workbook
from openpyxl.chart import (ScatterChart, Reference, Series)
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties


class LogPlotXlsx:
    """
    Plot the data as a xlsx format.
    """

    def __init__(self, plot_config, log_group_set):
        """
        Constructor.

        :param plot_config: Exported plot config from a LogAnalyzerConfig object
        :param log_group_set: LogGroupSet object
        """
        self.config = plot_config
        self.data_set = log_group_set
        self.workbook = Workbook()
        self.data_dict = None
        self.max_row = None
        self.data_worksheet = None

    def _setup(self):
        """
        Create xlsx file and tabs, export data matrix from the data set

        """
        # Add worksheets
        for i, group in enumerate(self.config):
            if i == 0:
                self.workbook.active.title = group.group_id
            else:
                self.workbook.create_sheet(group.group_id)
        self.workbook.create_sheet('CHART_DATA')
        self.data_worksheet = self.workbook['CHART_DATA']
        # Add chart data
        '''
        # Only add plotting data to the file
        data_series_set = set()
        for group in self.config:
            for plot in group.plot_list:
                if plot.y_axis.config_1 is not None:
                    data_series_set.update(plot.y_axis.config_1.data_series_list)
                if plot.y_axis.config_2 is not None:
                    data_series_set.update(plot.y_axis.config_2.data_series_list)
                if plot.x_axis.config_1 is not None:
                    data_series_set.update(plot.x_axis.config_1.data_series_list)
                if plot.x_axis.config_2 is not None:
                    data_series_set.update(plot.x_axis.config_2.data_series_list)
        data_mat, self.data_dict = self.data_set.export_matrix(list(data_series_set))
        '''
        # Add all data in the data set to the file
        data_mat, self.data_dict = self.data_set.export_matrix()
        self.max_row = len(data_mat)
        for row in data_mat:
            self.data_worksheet.append(row)

    def plot(self, path):
        """
        Plot the data to a xlsx file

        :param path: path to the xlsx file
        """
        self._setup()

        for group in self.config:
            worksheet = self.workbook[group.group_id]
            for i, plot in enumerate(group.plot_list): # Loop through each plot
                ''' Primary Plot'''
                cht1 = ScatterChart()
                # cht1.title = plot.title # Plot title

                '''Y1 Axis'''
                cht1.y_axis.title = plot.y_axis.config_1.label  # Label of the Y1 axis
                cht1.y_axis.scaling.min = plot.y_axis.config_1.min_val  # Max value on the Y1 axis
                cht1.y_axis.scaling.max = plot.y_axis.config_1.max_val  # Min value on the Y1 axis
                cht1.y_axis.majorUnit = plot.y_axis.config_1.major_tick  # MajorTick in the config file
                cht1.y_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=LineProperties(prstDash='solid')))
                cht1.y_axis.minorUnit = plot.y_axis.config_1.minor_tick  # MinorTick in the config file
                cht1.y_axis.minorGridlines = ChartLines(
                    spPr=GraphicalProperties(ln=LineProperties(prstDash='solid')))

                '''X Axis'''
                cht1.x_axis.title = plot.x_axis.config_1.label # Label of the X axis

                # Set plot min/max limits on X axis as the beginning/ending of the data if not specified in the config
                if plot.x_axis.config_1.min_val is not None:
                    cht1.x_axis.scaling.min = plot.x_axis.config_1.min_val
                else:
                    cht1.x_axis.scaling.min = self.data_worksheet.cell(
                                                    column=self.data_dict[plot.x_axis.config_1.data_series_list[0]] + 1,
                                                    row=2).value
                if plot.x_axis.config_1.max_val is not None:
                    cht1.x_axis.scaling.max = plot.x_axis.config_1.max_val
                else:
                    cht1.x_axis.scaling.max = self.data_worksheet.cell(
                                                    column=self.data_dict[plot.x_axis.config_1.data_series_list[0]] + 1,
                                                    row=self.max_row).value

                cht1.x_axis.majorUnit = plot.x_axis.config_1.major_tick
                cht1.x_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(ln=LineProperties(prstDash='solid')))
                cht1.x_axis.minorUnit = plot.x_axis.config_1.minor_tick
                cht1.x_axis.minorGridlines = ChartLines(
                    spPr=GraphicalProperties(ln=LineProperties(prstDash='solid')))
                cht1.x_axis.crosses = 'min'  # Place X-axis always on the lower value side
                cht1.legend.position = 'b'  # Place legend on the bottom

                '''Add data for Primary Plot'''
                for data_series in plot.y_axis.config_1.data_series_list:
                    data_col_y = self.data_dict[data_series] + 1 # Line # starts from 1
                    # Only support 1 data series on config_1 for X-axis
                    data_col_x = self.data_dict[plot.x_axis.config_1.data_series_list[0]] + 1 # Line # starts from 1
                    x_val = Reference(self.data_worksheet,
                                      min_col=data_col_x,
                                      min_row=2,
                                      max_row=self.max_row)
                    y_val = Reference(self.data_worksheet,
                                      min_col=data_col_y,
                                      min_row=1,
                                      max_row=self.max_row)
                    series = Series(values=y_val, xvalues=x_val, title_from_data=True)
                    cht1.series.append(series)

                ''' Secondary Plot '''
                if plot.y_axis.config_2 is not None:
                    cht2 = ScatterChart()
                    '''Y2 Axis'''
                    cht2.y_axis.title = plot.y_axis.config_2.label
                    cht2.y_axis.scaling.min = plot.y_axis.config_2.min_val
                    cht2.y_axis.scaling.max = plot.y_axis.config_2.max_val
                    cht2.y_axis.majorUnit = plot.y_axis.config_2.major_tick
                    cht2.y_axis.majorGridlines = ChartLines(
                        spPr=GraphicalProperties(ln=LineProperties(prstDash='lgDash')))
                    cht2.y_axis.minorUnit = plot.y_axis.config_2.minor_tick
                    cht2.y_axis.minorGridlines = ChartLines(
                        spPr=GraphicalProperties(ln=LineProperties(prstDash='dash')))
                    cht2.y_axis.axId = 200

                    '''Add data for Secondary Plot'''
                    for data_series in plot.y_axis.config_2.data_series_list:
                        data_col_y = self.data_dict[data_series] + 1
                        # Only support 1 data series on config_1 for X-axis
                        data_col_x = self.data_dict[plot.x_axis.config_1.data_series_list[0]] + 1
                        x_val = Reference(self.data_worksheet,
                                          min_col=data_col_x,
                                          min_row=2,
                                          max_row=self.max_row)
                        y_val = Reference(self.data_worksheet,
                                          min_col=data_col_y,
                                          min_row=1,
                                          max_row=self.max_row)
                        series = Series(values=y_val, xvalues=x_val, title_from_data=True)
                        cht2.series.append(series)
                    cht2.y_axis.crosses = 'max'
                    cht1 += cht2

                ''' Chart Size '''
                cht1.width = 25.4
                cht1.height = 10

                ''' Chart Position '''
                worksheet.add_chart(cht1, 'A' + str(i * 19 + 1))
                worksheet.cell(column=13,
                               row=i * 19 + 1,
                               value=plot.title)
        self.workbook.save(path)


class LogPlot:
    """
    Plot the data to various output formats.
    """

    def __init__(self, analyzer_config, log_group_set):
        self.config = analyzer_config.export_plot_config()
        self.data_set = log_group_set

    def plot_to_xlsx(self, path):
        """
        Function to output the data to xlsx format.

        :param path: output path to the xlsx file
        """
        xlsx_plot = LogPlotXlsx(self.config, self.data_set)
        xlsx_plot.plot(path)
